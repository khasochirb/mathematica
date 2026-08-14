#!/usr/bin/env python3
"""Mongol Potential — centre + website pricing model.

Builds a live-formula .xlsx. Every output traces back to the Assumptions sheet.
No price is recommended; the model produces ranges and break-evens.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- style
FONT = "Arial"
H1 = Font(name=FONT, size=14, bold=True, color="1B2430")
H2 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
H3 = Font(name=FONT, size=10, bold=True, color="1B2430")
BODY = Font(name=FONT, size=10)
SMALL = Font(name=FONT, size=9, color="595959")
INPUT = Font(name=FONT, size=10, color="0000FF")          # hardcoded input
CALC = Font(name=FONT, size=10)                            # formula
LINK = Font(name=FONT, size=10, color="008000")            # cross-sheet link
WARN = Font(name=FONT, size=10, bold=True, color="C00000")

HDR_FILL = PatternFill("solid", fgColor="1B2430")
KEY_FILL = PatternFill("solid", fgColor="FFFF00")          # user should edit
SEC_FILL = PatternFill("solid", fgColor="E8EAF6")
GOOD_FILL = PatternFill("solid", fgColor="E8F5E9")
BAD_FILL = PatternFill("solid", fgColor="FDECEA")

MNT = '#,##0;(#,##0);-'
MNT0 = '#,##0;(#,##0);-'
PCT = '0.0%'
NUM1 = '0.0'
NUM0 = '0'
USD = '$#,##0'

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()


def sheet(name, widths):
    ws = wb.create_sheet(name)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    return ws


def title(ws, row, text, span=6):
    c = ws.cell(row=row, column=1, value=text)
    c.font = H1
    return row + 1


def section(ws, row, text, span=6):
    for col in range(1, span + 1):
        cc = ws.cell(row=row, column=col)
        cc.fill = HDR_FILL
        cc.font = H2
    ws.cell(row=row, column=1, value=text)
    return row + 1


def note(ws, row, text, col=1, span=8):
    c = ws.cell(row=row, column=col, value=text)
    c.font = SMALL
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


# ================================================================ README
ws = sheet("README", [3, 92, 18, 18, 18, 18])
r = 1
r = title(ws, r, "Mongol Potential — Pricing & Unit Economics Model")
r = note(ws, r, "Built 14 Aug 2026 for the centre opening 1 Sept 2026. Currency: Mongolian tugrik (MNT, ₮) unless marked USD.", col=2)
r += 1

r = section(ws, r, "  WHAT THIS IS — AND WHAT IT IS NOT")
for t in [
    "This model gives you the ARITHMETIC to choose a price. It deliberately does NOT recommend one.",
    "Every table shows a RANGE of price points and what each one implies for break-even, profit and risk. The choice is yours.",
    "",
    "Change the numbers on the ASSUMPTIONS sheet. Everything else recalculates. Do not type over formulas (black text).",
    "  •  BLUE text  = an input you can edit",
    "  •  YELLOW fill = a key assumption most worth stress-testing",
    "  •  BLACK text = a formula, do not overwrite",
    "  •  GREEN text = a link to another sheet",
]:
    r = note(ws, r, t, col=2)
r += 1

r = section(ws, r, "  SHEET GUIDE")
for name, desc in [
    ("Assumptions", "Every editable input. Start here."),
    ("Benchmarks", "Real Ulaanbaatar market prices found, with sources and dates. Also lists what could NOT be verified."),
    ("Capacity", "Rooms → sessions per week → realistic student capacity. Utilisation, not theoretical maximum."),
    ("Costs", "Monthly cost base. Includes owner-taught vs all-hired staffing comparison."),
    ("PriceGrid", "The core output: break-even student count and profit at each price point × cohort size."),
    ("PnL", "Full monthly and annual P&L at the price you select on Assumptions."),
    ("Enrolment", "Full-school-year enrolment vs monthly subscription, including the cost of churn."),
    ("Website", "Bundled into the centre fee vs sold as an add-on. Then public (non-centre) website pricing."),
    ("CashFlow", "Month-by-month Aug 2026 → Aug 2027 against the school year. Tight months flagged."),
    ("Scenarios", "Conservative / Base / Stretch side by side."),
]:
    ws.cell(row=r, column=2, value=name).font = H3
    ws.cell(row=r, column=3, value=desc).font = BODY
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    r += 1
r += 1

r = section(ws, r, "  THREE THINGS TO KNOW BEFORE YOU READ THE NUMBERS")
for t in [
    "1.  MISSING SOURCE DOCUMENTS.  The brief asked me to read /docs/plan/03-ROADMAP.md and /docs/plan/05-INTEGRITY-AND-MOTIVATION.md.",
    "     Neither file exists in this repository or anywhere on this machine (the repo has only docs/superpowers/). I have NOT guessed at their",
    "     contents. The premium-justification logic on the Website sheet is built from the mechanism described in the brief itself",
    "     (verified in-room attempts → trustworthy predicted ЭЕШ score → monthly parent report). Re-check that section against doc 05.",
    "",
    "2.  THE OWNER'S TEACHING TIME IS COSTED, not free. See Assumptions rows 'Owner teaching'. If you zero it out the model will",
    "     flatter itself and mislead the pricing decision. Keep it costed even while you are not paying yourself.",
    "",
    "3.  MARKET DATA IS THIN AND PARTLY UNVERIFIABLE. Mongolian tutoring centres mostly publish prices on Facebook, not on indexable",
    "     websites, and several .mn domains could not be opened from this machine. Benchmarks marked LOW confidence are search-snippet",
    "     level only. Before you commit to a price, phone three competitors and replace those cells with what they actually quote.",
]:
    r = note(ws, r, t, col=2)

# ============================================================ ASSUMPTIONS
ws = sheet("Assumptions", [3, 52, 16, 12, 62])
A = {}          # key -> row
r = 1
r = title(ws, r, "Assumptions — edit these")
r = note(ws, r, "Blue = editable input. Yellow = key lever. Everything on other sheets flows from here.", col=2)
r += 1


def head(row, label):
    for col in range(1, 6):
        cc = ws.cell(row=row, column=col)
        cc.fill = HDR_FILL
        cc.font = H2
    ws.cell(row=row, column=1, value=label)
    return row + 1


def put(row, key, label, value, fmt, src="", key_lever=False, unit=""):
    ws.cell(row=row, column=2, value=label).font = BODY
    c = ws.cell(row=row, column=3, value=value)
    c.font = INPUT
    c.number_format = fmt
    c.border = BOX
    if key_lever:
        c.fill = KEY_FILL
    ws.cell(row=row, column=4, value=unit).font = SMALL
    ws.cell(row=row, column=5, value=src).font = SMALL
    ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    A[key] = row
    return row + 1


r = head(r, "  GENERAL")
r = put(r, "fx", "Exchange rate (MNT per USD)", 3599, MNT, "xe.com, 13 Aug 2026: 1 USD = 3,599.11 MNT", unit="₮/$")
r = put(r, "months", "Billing months per school year", 10, NUM0, "Sept–June. Change if you bill 9 or 12.", unit="months", key_lever=True)
r = put(r, "weeks", "Weeks per month (teaching)", 4.33, NUM1, "52/12", unit="weeks")

r = head(r, "  CAPACITY — rooms and sessions")
r = put(r, "capA", "Room A (living room) — max students", 20, NUM0, "Brief: '~20 students per 2-hour session'. Treated as the main teaching room.", unit="students", key_lever=True)
r = put(r, "capB", "Room B (small bedroom) — max students", 8, NUM0, "ASSUMPTION — not stated in brief. Set to 0 to model one room only.", unit="students", key_lever=True)
r = put(r, "useB", "Use Room B in parallel? (1 = yes, 0 = no)", 1, NUM0, "Parallel sessions need a second teacher at the same time.", unit="1/0")
r = put(r, "concur_cap", "Centre-wide concurrent cap (fire/comfort)", 20, NUM0, "Warning fires on Capacity sheet if cohorts exceed this.", unit="students")
r = put(r, "sess_hours", "Session length", 2, NUM1, "Brief.", unit="hours")
r = put(r, "wd_slots", "Weekday after-school slots per day per room", 2, NUM0, "e.g. 16:00–18:00 and 18:00–20:00.", unit="slots")
r = put(r, "wd_days", "Weekdays used", 5, NUM0, "", unit="days")
r = put(r, "we_slots", "Weekend slots per day per room", 4, NUM0, "e.g. 10:00–18:00.", unit="slots")
r = put(r, "we_days", "Weekend days used", 2, NUM0, "", unit="days")
r = put(r, "util", "Realistic slot utilisation", 0.55, PCT, "NOT theoretical max. Demand concentrates in the evening peak; some slots will never fill.", unit="%", key_lever=True)
r = put(r, "sess_wk", "Sessions per student per week", 2, NUM0, "Benchmark: Eduzone ЭЕШ maths runs 2×/week.", unit="sessions", key_lever=True)

r = head(r, "  ENROLMENT")
r = put(r, "cohort", "Cohort size — selected scenario", 16, NUM0, "PriceGrid also shows 12 / 16 / 20 side by side.", unit="students", key_lever=True)
r = put(r, "cohortB", "Cohort size — Room B small group", 6, NUM0, "ASSUMPTION.", unit="students")
r = put(r, "fill", "Seat fill rate", 0.85, PCT, "Share of available seats actually sold.", unit="%", key_lever=True)
r = put(r, "yr_share", "Share enrolling for the full school year", 0.60, PCT, "Owner prefers full-year; monthly accepted. ASSUMPTION.", unit="%", key_lever=True)
r = put(r, "churn", "Monthly churn — monthly-plan students", 0.05, PCT, "ASSUMPTION. Stress-test this: it drives the year-vs-monthly answer.", unit="%/mo", key_lever=True)

r = head(r, "  PRICE (the decision — model shows a range, this is just the selected point)")
r = put(r, "price", "Tuition per student per month", 450000, MNT, "SELECTED SCENARIO ONLY. See PriceGrid for the full range 200k–900k.", unit="₮/mo", key_lever=True)
r = put(r, "yr_disc", "Discount for paying full year upfront", 0.10, PCT, "ASSUMPTION.", unit="%", key_lever=True)
r = put(r, "reg_fee", "One-off registration / materials fee", 0, MNT, "Set to 0 if not charged.", unit="₮")

r = head(r, "  STAFF COSTS — owner time IS costed (see README point 2)")
r = put(r, "teach_sal", "Hired teacher — monthly gross salary", 3500000, MNT, "Mongolian teacher BASIC salary rose to 2.8m₮/mo Jan 2026 (+50%), +26% more due 1 Nov 2026, union target 3.5m₮ (Education International). A strong private-centre maths teacher costs a premium over basic.", unit="₮/mo", key_lever=True)
r = put(r, "teach_hrs", "Hired teacher — contact hours capacity per week", 20, NUM0, "Teaching hours only; excludes prep/marking.", unit="hrs/wk")
r = put(r, "si", "Employer social insurance", 0.145, PCT, "Mongolia employer contribution 12.5%–14.5%; TradingEconomics shows 14.5% company rate.", unit="%")
r = put(r, "own_hrs", "Owner teaching hours per week", 12, NUM1, "Contact hours the owner personally teaches.", unit="hrs/wk", key_lever=True)
r = put(r, "own_rate", "Owner teaching — cost per hour (opportunity cost)", 50000, MNT, "MUST NOT be zero. Proxy: private lessons in UB advertise from ₮50,000/hr (TUTOROO) — that listing is LANGUAGE tutoring, so treat as a floor, not a maths rate.", unit="₮/hr", key_lever=True)
r = put(r, "admin_sal", "Admin / front desk — monthly gross", 1200000, MNT, "ASSUMPTION. Set headcount to 0 if the owner does admin (then raise owner hours).", unit="₮/mo")
r = put(r, "admin_n", "Admin headcount", 1, NUM1, "", unit="FTE")

r = head(r, "  FACILITY")
r = put(r, "rent", "Rent, or imputed market rent if owned", 2000000, MNT, "Cost it even if you own it. UB 2-room average rent ₮1.85m/mo (ikon.mn, Feb 2026); Sukhbaatar district 1.40–4.61m. A ground-floor unit fit for a centre likely sits at/above the top of the 2-room range.", unit="₮/mo", key_lever=True)
r = put(r, "util_cost", "Utilities (heating, power, water)", 400000, MNT, "UNVERIFIED — no source found. Replace with your actual bill.", unit="₮/mo")
r = put(r, "internet", "Internet", 60000, MNT, "UNVERIFIED — replace with actual.", unit="₮/mo")

r = head(r, "  VARIABLE COST PER STUDENT")
r = put(r, "materials", "Materials & printing per student", 15000, MNT, "UNVERIFIED — replace with actual.", unit="₮/student/mo")
r = put(r, "food", "Food & coffee per student (under consideration)", 25000, MNT, "UNVERIFIED. Set to 0 to test dropping the provision.", unit="₮/student/mo", key_lever=True)

r = head(r, "  TECHNOLOGY & OVERHEAD")
r = put(r, "supabase", "Supabase Pro", 25, USD, "Published Supabase Pro price.", unit="$/mo")
r = put(r, "hosting", "Vercel, domain, email, misc SaaS", 25, USD, "ASSUMPTION.", unit="$/mo")
r = put(r, "marketing", "Marketing", 500000, MNT, "ASSUMPTION.", unit="₮/mo", key_lever=True)
r = put(r, "other", "Other / contingency", 300000, MNT, "ASSUMPTION.", unit="₮/mo")
r = put(r, "cac", "Cost to acquire one replacement student", 250000, MNT, "ASSUMPTION. Used to price churn on the Enrolment sheet.", unit="₮", key_lever=True)
r = put(r, "pay_summer", "Pay hired teachers through the summer? (1 = yes)", 1, NUM0, "Set to 1 if you keep staff on annual contracts (you usually must, to have them in September). Owner teaching time is never charged in a non-teaching month.", unit="1/0", key_lever=True)

r = head(r, "  WEBSITE")
r = put(r, "web_addon", "Add-on price to centre students (if NOT bundled)", 50000, MNT, "ASSUMPTION.", unit="₮/mo", key_lever=True)
r = put(r, "web_attach", "Add-on attach rate (if sold separately)", 0.40, PCT, "ASSUMPTION — share of centre students who buy it.", unit="%", key_lever=True)
r = put(r, "bundle_uplift", "Tuition uplift the bundle supports (if bundled)", 0.08, PCT, "The report makes the fee legible, so the fee can be higher. ASSUMPTION — this is the crux of the bundling case.", unit="%", key_lever=True)
r = put(r, "bundle_churn", "Churn reduction from bundling", 0.015, PCT, "Percentage POINTS off monthly churn. ASSUMPTION.", unit="pp", key_lever=True)
r = put(r, "pub_price", "Public (non-centre) paid tier", 30000, MNT, "ASSUMPTION.", unit="₮/mo", key_lever=True)
r = put(r, "pub_users", "Public registered users today", 12, NUM0, "Stated in brief.", unit="users")
r = put(r, "pub_growth", "Public registered users added per month", 150, NUM0, "ASSUMPTION — no traffic data supplied.", unit="users/mo", key_lever=True)
r = put(r, "pub_conv", "Free → paid conversion", 0.02, PCT, "ASSUMPTION. 2% is a common freemium benchmark; unvalidated here.", unit="%", key_lever=True)

ASU = "Assumptions!$C$"


def a(key):
    return f"{ASU}{A[key]}"


# ============================================================== BENCHMARKS
ws = sheet("Benchmarks", [3, 34, 26, 15, 13, 46])
r = 1
r = title(ws, r, "Ulaanbaatar market benchmarks")
r = note(ws, r, "Everything below was found by web search on 14 Aug 2026. Confidence is graded honestly. LOW = search-snippet only, primary page could not be opened from this machine (many .mn domains were blocked). Verify by phone before pricing.", col=2)
r += 1

hdr = ["", "What", "Price", "Basis", "Confidence", "Source / date"]
for i, h in enumerate(hdr, start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.fill = HDR_FILL
    c.font = H2
r += 1

rows = [
    ("GROUP ЭЕШ PREP — the market you are pricing against", "", "", "", ""),
    ("Eduzone сургалтын төв — maths ЭЕШ prep", "1,800,000 ₮ / year", "2 sessions/week, group", "MEDIUM", "unread.today course listing, via search 14 Aug 2026"),
    ("Eduzone — physics / English", "1,600,000 ₮ / year", "2 sessions/week, group", "MEDIUM", "same listing"),
    ("iMath.mn — summer programme", "360,000 ₮", "daily 15:00–17:00, summer block", "LOW", "search snippet 14 Aug 2026; imath.mn blocked from this machine"),
    ("Оюунлаг school — бэлтгэл анги (prep class)", "900,000 ₮", "prep class, period not stated", "LOW", "shuurhai.mn private-school roundup, via search"),
    ("", "", "", "", ""),
    ("PRIVATE 1-ON-1 TUTORING", "", "", "", ""),
    ("UB private lessons (TUTOROO listings)", "from 50,000 ₮ / hour", "1-on-1, in person/online", "LOW — WRONG SUBJECT", "tutoroo.co, Mongolian- and English-language tutors. NOT maths. Use as a floor only."),
    ("Online Mongolian-language lessons", "$10–30 / hour (avg $18)", "1-on-1 online", "LOW — WRONG SUBJECT", "AmazingTalker / Superprof. Language, not maths."),
    ("", "", "", "", ""),
    ("SAT PREP", "", "", "", ""),
    ("SAT prep course price in Mongolia", "NOT FOUND", "—", "GAP", "No published price located. Thin market, as expected. Phone EA MUST / GAPE / Absolute."),
    ("IELTS exam fee (adjacent test-prep anchor)", "839,000 ₮", "exam sitting, not tuition", "MEDIUM", "ikon.mn 2025 IELTS schedule"),
    ("", "", "", "", ""),
    ("TOP OF MARKET — what UB parents already pay for perceived quality", "", "", "", ""),
    ("Private secondary schools — average", "8,000,000–10,000,000 ₮ / year", "full school", "MEDIUM", "eagle.mn"),
    ("Private secondary schools — full range", "1,700,000–59,000,000 ₮ / year", "full school", "MEDIUM", "ikon.mn"),
    ("Оюунлаг school", "8,512,000 ₮ / year", "full school", "MEDIUM", "shuurhai.mn roundup"),
    ("Улаанбаатар Элит school", "9,950,000 ₮ / year", "full school", "MEDIUM", "shuurhai.mn roundup"),
    ("International School of Ulaanbaatar", "$24,190–$44,060 / year", "full school, no employer aid", "HIGH", "ISU published tuition-fees PDF, 2025/26"),
    ("", "", "", "", ""),
    ("COST INPUTS", "", "", "", ""),
    ("Teacher basic salary — from Jan 2026", "2,800,000 ₮ / month", "state basic, +50%", "HIGH", "Education International, Dec 2025 settlement"),
    ("Further rise due 1 Nov 2026", "+26%", "on basic", "HIGH", "Education International"),
    ("Union target", "3,500,000 ₮ / month", "on basic", "HIGH", "Education International / FMESU"),
    ("Employer social insurance", "12.5%–14.5%", "on gross payroll", "HIGH", "PwC / TradingEconomics (14.5% company rate)"),
    ("UB 2-room apartment rent — average", "1,850,000 ₮ / month", "residential", "MEDIUM", "ikon.mn, Feb 2026"),
    ("UB 2-room rent — Sukhbaatar district", "1,400,000–4,610,000 ₮ / month", "residential", "MEDIUM", "ikon.mn"),
    ("Exchange rate", "1 USD = 3,599.11 ₮", "spot", "HIGH", "xe.com, 13 Aug 2026"),
]
for label, price, basis, conf, src in rows:
    if price == "" and basis == "":
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = SEC_FILL
        c = ws.cell(row=r, column=2, value=label)
        c.font = H3
    else:
        ws.cell(row=r, column=2, value=label).font = BODY
        ws.cell(row=r, column=3, value=price).font = BODY
        ws.cell(row=r, column=4, value=basis).font = SMALL
        cc = ws.cell(row=r, column=5, value=conf)
        cc.font = Font(name=FONT, size=9, bold=True,
                       color="C00000" if conf.startswith(("LOW", "GAP")) else "1B5E20")
        ws.cell(row=r, column=6, value=src).font = SMALL
        ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = section(ws, r, "  WHAT I COULD NOT FIND — do not let these gaps get filled with guesses")
for t in [
    "•  A published per-hour rate for private MATHS tutoring in Ulaanbaatar. The ₮50,000/hr figure is language tutoring on an expat-facing platform.",
    "•  Any SAT preparation course price in Mongolia.",
    "•  Monthly (as opposed to annual) fees for ЭЕШ centres — most advertise a year price or quote by phone.",
    "•  Utilities and heating cost for a converted ground-floor apartment.",
    "•  Competitor cohort sizes, which is what actually determines their margin and therefore how much room you have above them.",
    "",
    "Most Mongolian centres price on Facebook and Instagram, not on indexable websites, and several .mn domains were blocked from this machine.",
    "The fastest fix is three phone calls. Ask: price, sessions per week, group size, and whether that price is per month or per year.",
]:
    r = note(ws, r, t, col=2)

# ================================================================ CAPACITY
ws = sheet("Capacity", [3, 50, 16, 14, 60])
C = {}
r = 1
r = title(ws, r, "Capacity — realistic, not theoretical")
r = note(ws, r, "A 'slot' is one 2-hour session in one room. A cohort meets 'sessions per week' times, so each cohort consumes that many slots.", col=2)
r += 1


def calc(row, key, label, formula, fmt, src="", store=None, font=CALC):
    ws.cell(row=row, column=2, value=label).font = BODY
    c = ws.cell(row=row, column=3, value=formula)
    c.font = font
    c.number_format = fmt
    c.border = BOX
    ws.cell(row=row, column=5, value=src).font = SMALL
    ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    if store is not None:
        store[key] = row
    return row + 1


r = section(ws, r, "  SLOTS PER WEEK, PER ROOM")
r = calc(r, "theo", "Theoretical slots per week per room",
         f"=({a('wd_slots')}*{a('wd_days')})+({a('we_slots')}*{a('we_days')})", NUM1,
         "Every slot the timetable could physically hold.", C, LINK)
r = calc(r, "real", "Realistic slots per week per room (after utilisation)",
         f"=C{C['theo']}*{a('util')}", NUM1,
         "This is the number the model uses. The evening peak fills; mid-afternoon largely does not.", C)
r = calc(r, "slots_tot", "Realistic slots per week — whole centre",
         f"=C{C['real']}*(1+{a('useB')})", NUM1, "Room A, plus Room B if in use.", C)

r += 1
r = section(ws, r, "  COHORTS AND STUDENTS")
r = calc(r, "cohA", "Cohorts that fit in Room A",
         f"=C{C['real']}/{a('sess_wk')}", NUM1, "Each cohort uses 'sessions per week' slots.", C)
r = calc(r, "cohB", "Cohorts that fit in Room B",
         f"=C{C['real']}/{a('sess_wk')}*{a('useB')}", NUM1, "", C)
r = calc(r, "stuA", "Students — Room A",
         f"=C{C['cohA']}*MIN({a('cohort')},{a('capA')})*{a('fill')}", NUM1,
         "Cohort size is capped by the room's physical capacity.", C)
r = calc(r, "stuB", "Students — Room B",
         f"=C{C['cohB']}*MIN({a('cohortB')},{a('capB')})*{a('fill')}", NUM1, "", C)
r = calc(r, "students", "TOTAL STUDENTS AT STEADY STATE",
         f"=ROUND(C{C['stuA']}+C{C['stuB']},0)", NUM0,
         "The enrolled roll the centre can actually serve on these assumptions.", C)
ws.cell(row=C["students"], column=3).font = Font(name=FONT, size=11, bold=True)
ws.cell(row=C["students"], column=3).fill = GOOD_FILL

r += 1
r = section(ws, r, "  TEACHING LOAD")
r = calc(r, "hours", "Contact hours per week (all cohorts)",
         f"=C{C['slots_tot']}*{a('sess_hours')}", NUM1, "", C)
r = calc(r, "own_h", "…covered by the owner",
         f"=MIN({a('own_hrs')},C{C['hours']})", NUM1, "", C)
r = calc(r, "hire_h", "…left for hired teachers",
         f"=MAX(0,C{C['hours']}-C{C['own_h']})", NUM1, "", C)
r = calc(r, "teachers", "Hired teachers required",
         f"=ROUNDUP(C{C['hire_h']}/{a('teach_hrs')},0)", NUM0,
         "Rounded up — you cannot hire a fraction of a teacher.", C)
ws.cell(row=C["teachers"], column=3).font = Font(name=FONT, size=11, bold=True)

r += 1
r = section(ws, r, "  CHECKS")
r = calc(r, "concur", "Students in the building during a parallel session",
         f"=MIN({a('cohort')},{a('capA')})+MIN({a('cohortB')},{a('capB')})*{a('useB')}", NUM0, "", C)
ws.cell(row=r, column=2, value="Within the centre-wide concurrent cap?").font = BODY
cc = ws.cell(row=r, column=3, value=f'=IF(C{C["concur"]}<={a("concur_cap")},"OK","OVER CAP — reduce cohort size or stop using Room B in parallel")')
cc.font = WARN
cc.border = BOX
r += 1
ws.cell(row=r, column=2, value="Second teacher needed at the same time?").font = BODY
cc = ws.cell(row=r, column=3, value=f'=IF({a("useB")}=1,"YES — parallel sessions need 2 adults in the building","No")')
cc.font = BODY
cc.border = BOX
r += 2
r = note(ws, r, "If 'OVER CAP' appears: either the cohort is too big for the flat, or Room B should run at a different time rather than in parallel. Running Room B in parallel is what forces a second salaried teacher — check on the Costs sheet what that costs you.", col=2)
r += 1
r = note(ws, r, "IMPORTANT — this is the STEADY-STATE roll, not what you will have in September. It is the number of enrolled students the timetable can serve once the centre is full. The CashFlow sheet ramps up to it. A new centre does not open full.", col=2)
r += 1
r = calc(r, "phys", "Physical ceiling on the roll (100% fill, for reference)",
         f"=(C{C['real']}*{a('capA')}+C{C['real']}*{a('capB')}*{a('useB')})/{a('sess_wk')}", NUM0,
         "Every seat in every realistic slot, sold. The roll above should sit comfortably below this.", C)

CAP = "Capacity!$C$"

# =================================================================== COSTS
ws = sheet("Costs", [3, 50, 16, 14, 60])
K = {}
r = 1
r = title(ws, r, "Monthly cost base")
r = note(ws, r, "Fixed costs do not move with student numbers. Variable costs do. The split is what sets the break-even.", col=2)
r += 1

r = section(ws, r, "  FIXED COSTS PER MONTH")
r = calc(r, "c_teach", "Hired teachers (incl. employer social insurance)",
         f"={CAP}{C['teachers']}*{a('teach_sal')}*(1+{a('si')})", MNT,
         "Headcount from Capacity sheet.", K, LINK)
r = calc(r, "c_owner", "Owner teaching time (opportunity cost — NOT free)",
         f"={CAP}{C['own_h']}*{a('weeks')}*{a('own_rate')}", MNT,
         "Costed even though unpaid today. Removing this would flatter the model and mislead the price decision.", K, LINK)
ws.cell(row=K["c_owner"], column=3).fill = KEY_FILL
r = calc(r, "c_admin", "Admin / front desk (incl. social insurance)",
         f"={a('admin_n')}*{a('admin_sal')}*(1+{a('si')})", MNT, "", K)
r = calc(r, "c_rent", "Rent (or imputed rent if owned)",
         f"={a('rent')}", MNT, "", K)
r = calc(r, "c_util", "Utilities + internet",
         f"={a('util_cost')}+{a('internet')}", MNT, "UNVERIFIED inputs.", K)
r = calc(r, "c_tech", "Technology (Supabase, hosting, domain)",
         f"=({a('supabase')}+{a('hosting')})*{a('fx')}", MNT, "Converted at the rate on Assumptions.", K)
r = calc(r, "c_mkt", "Marketing", f"={a('marketing')}", MNT, "", K)
r = calc(r, "c_oth", "Other / contingency", f"={a('other')}", MNT, "", K)
r = calc(r, "fixed", "TOTAL FIXED COST PER MONTH",
         f"=SUM(C{K['c_teach']}:C{K['c_oth']})", MNT, "", K)
ws.cell(row=K["fixed"], column=3).font = Font(name=FONT, size=11, bold=True)
ws.cell(row=K["fixed"], column=3).fill = SEC_FILL

r += 1
r = section(ws, r, "  VARIABLE COST PER STUDENT PER MONTH")
r = calc(r, "v_mat", "Materials & printing", f"={a('materials')}", MNT, "", K)
r = calc(r, "v_food", "Food & coffee", f"={a('food')}", MNT, "Set to 0 on Assumptions to test dropping it.", K)
r = calc(r, "var", "TOTAL VARIABLE COST PER STUDENT",
         f"=C{K['v_mat']}+C{K['v_food']}", MNT, "", K)
ws.cell(row=K["var"], column=3).font = Font(name=FONT, size=11, bold=True)
ws.cell(row=K["var"], column=3).fill = SEC_FILL

r += 1
r = section(ws, r, "  STAFFING MODEL — owner-taught vs all-hired")
r = note(ws, r, "Same teaching load, two ways to cover it. The difference is what the owner's time is worth.", col=2)
r += 1
ws.cell(row=r, column=2, value="Model A — owner teaches (current assumption)").font = H3
r += 1
r = calc(r, "A_teach", "  Hired teacher cost", f"=C{K['c_teach']}", MNT, "", K)
r = calc(r, "A_own", "  Owner time costed", f"=C{K['c_owner']}", MNT, "", K)
r = calc(r, "A_tot", "  Total teaching cost", f"=C{K['A_teach']}+C{K['A_own']}", MNT, "", K)
r += 1
ws.cell(row=r, column=2, value="Model B — all teaching hired, owner teaches nothing").font = H3
r += 1
r = calc(r, "B_n", "  Hired teachers required",
         f"=ROUNDUP({CAP}{C['hours']}/{a('teach_hrs')},0)", NUM0, "", K, LINK)
r = calc(r, "B_teach", "  Hired teacher cost",
         f"=C{K['B_n']}*{a('teach_sal')}*(1+{a('si')})", MNT, "", K)
r = calc(r, "B_tot", "  Total teaching cost", f"=C{K['B_teach']}", MNT, "", K)
r += 1
r = calc(r, "delta", "Cash saved per month by the owner teaching",
         f"=C{K['B_tot']}-C{K['A_teach']}", MNT,
         "Real cash difference. Compare against the owner's costed time directly below — if the owner's time is worth more than this, hiring is the better trade.", K)
r = calc(r, "delta2", "…versus the value of the owner's time",
         f"=C{K['c_owner']}", MNT, "", K)
ws.cell(row=r, column=2, value="Owner teaching is economically worth it?").font = BODY
cc = ws.cell(row=r, column=3, value=f'=IF(C{K["delta"]}>=C{K["c_owner"]},"YES — saves more cash than the time is worth","NO — the owner\'s time is worth more than the saving")')
cc.font = WARN
cc.border = BOX
r += 2
r = note(ws, r, "This comparison is only as honest as the owner hourly rate on Assumptions. If that rate is set too low, owner-teaching will always look free and the centre will look more profitable than it is.", col=2)

KO = "Costs!$C$"

# =============================================================== PRICEGRID
ws = sheet("PriceGrid", [3, 20, 18, 20, 18, 18, 18, 18, 18, 18])
r = 1
r = title(ws, r, "Price grid — the core output")
r = note(ws, r, "No price is recommended. Read across: what each price implies for break-even and profit, at three cohort sizes. Break-even does NOT change with cohort size (the same rooms, teachers and rent) — cohort size determines whether you can REACH it.", col=2)
r += 1

ws.cell(row=r, column=2, value="Seats available at each cohort size (steady state):").font = H3
for j, coh in enumerate([12, 16, 20]):
    ws.cell(row=r, column=6 + j, value=f"=ROUND({CAP}{C['cohA']}*MIN({coh},{a('capA')})*{a('fill')}+{CAP}{C['cohB']}*MIN({a('cohortB')},{a('capB')})*{a('fill')},0)").number_format = NUM0
    ws.cell(row=r, column=6 + j).font = LINK
    ws.cell(row=r, column=6 + j).fill = SEC_FILL
    ws.cell(row=r, column=6 + j).border = BOX
r += 1
r = note(ws, r, "Compare these against the break-even column: if break-even exceeds the seats available, that price cannot work at this capacity.", col=2)
r += 1

price_hdr_row = r
headers = ["", "Tuition\n₮/student/month", "Annual\n₮/student", "vs Eduzone ЭЕШ\n(1.8m₮/yr)", "Break-even\nstudents",
           "Profit/month\ncohort 12", "Profit/month\ncohort 16", "Profit/month\ncohort 20",
           "Profit/year\ncohort 16", "Margin\ncohort 16"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.fill = HDR_FILL
    c.font = H2
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
ws.row_dimensions[r].height = 42
r += 1

first_data = r
prices = [200000, 250000, 300000, 350000, 400000, 450000, 500000, 550000,
          600000, 650000, 700000, 750000, 800000, 900000]

# students at a given cohort size = cohortsA*MIN(coh,capA)*fill + cohortsB*MIN(cohB,capB)*fill
def students_expr(coh_literal):
    return (f"({CAP}{C['cohA']}*MIN({coh_literal},{a('capA')})*{a('fill')}"
            f"+{CAP}{C['cohB']}*MIN({a('cohortB')},{a('capB')})*{a('fill')})")


for p in prices:
    ws.cell(row=r, column=2, value=p).font = INPUT
    ws.cell(row=r, column=2).number_format = MNT
    ws.cell(row=r, column=2).border = BOX
    # annual
    ws.cell(row=r, column=3, value=f"=B{r}*{a('months')}").number_format = MNT
    ws.cell(row=r, column=3).font = CALC
    # vs eduzone benchmark 1,800,000
    ws.cell(row=r, column=4, value=f"=C{r}/1800000").number_format = '0.0"×"'
    ws.cell(row=r, column=4).font = CALC
    # effective price after full-year discount blend
    eff = f"(B{r}*(1-{a('yr_share')}*{a('yr_disc')}))"
    # break-even students = fixed / (eff - var)
    ws.cell(row=r, column=5, value=f"=IFERROR({KO}{K['fixed']}/({eff}-{KO}{K['var']}),\"n/a\")").number_format = NUM0
    ws.cell(row=r, column=5).font = CALC
    # profit at cohort 12/16/20
    for j, coh in enumerate([12, 16, 20]):
        col = 6 + j
        s = students_expr(coh)
        ws.cell(row=r, column=col,
                value=f"={s}*({eff}-{KO}{K['var']})-{KO}{K['fixed']}").number_format = MNT
        ws.cell(row=r, column=col).font = CALC
    # annual profit cohort 16
    ws.cell(row=r, column=9, value=f"=G{r}*{a('months')}").number_format = MNT
    ws.cell(row=r, column=9).font = CALC
    # margin cohort 16
    s16 = students_expr(16)
    ws.cell(row=r, column=10, value=f"=IFERROR(G{r}/({s16}*{eff}),\"n/a\")").number_format = PCT
    ws.cell(row=r, column=10).font = CALC
    for col in range(2, 11):
        ws.cell(row=r, column=col).border = BOX
    r += 1
last_data = r - 1

r += 1
r = note(ws, r, "Column D anchors you to the only solid competitor price found: Eduzone charges 1,800,000₮/year for 2×/week group maths ЭЕШ prep. A '3.0×' means you are asking parents for three times what the nearest visible competitor asks. That is a defensible position only if the difference is visible to them every month — which is what the parent report is for (see Website sheet).", col=2)
r += 1
r = note(ws, r, "Seats available: at cohort 12 / 16 / 20 the centre can enrol the number of students shown on the Capacity sheet. If break-even students exceeds seats available at your chosen price, that price cannot work at this capacity — raise price, raise utilisation, or cut fixed cost.", col=2)

from openpyxl.formatting.rule import CellIsRule
red = PatternFill("solid", fgColor="FDECEA")
green = PatternFill("solid", fgColor="E8F5E9")
for col in "FGHI":
    ws.conditional_formatting.add(f"{col}{first_data}:{col}{last_data}",
                                  CellIsRule(operator="lessThan", formula=["0"], fill=red))
    ws.conditional_formatting.add(f"{col}{first_data}:{col}{last_data}",
                                  CellIsRule(operator="greaterThan", formula=["0"], fill=green))

# ===================================================================== PnL
ws = sheet("PnL", [3, 50, 18, 18, 60])
P = {}
r = 1
r = title(ws, r, "P&L at the selected price")
r = note(ws, r, "Driven by the price on the Assumptions sheet. Change it there and this recalculates.", col=2)
r += 1

ws.cell(row=r, column=3, value="Per month").font = H3
ws.cell(row=r, column=4, value="Per year").font = H3
r += 1

r = section(ws, r, "  REVENUE")


def pnl(row, key, label, m_formula, y_formula, fmt=MNT, src="", bold=False, font=CALC):
    ws.cell(row=row, column=2, value=label).font = H3 if bold else BODY
    c = ws.cell(row=row, column=3, value=m_formula)
    c.font = Font(name=FONT, size=10, bold=True) if bold else font
    c.number_format = fmt
    c.border = BOX
    c2 = ws.cell(row=row, column=4, value=y_formula)
    c2.font = Font(name=FONT, size=10, bold=True) if bold else font
    c2.number_format = fmt
    c2.border = BOX
    ws.cell(row=row, column=5, value=src).font = SMALL
    ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    P[key] = row
    return row + 1


r = pnl(r, "students", "Students enrolled", f"={CAP}{C['students']}", f"={CAP}{C['students']}", NUM0, "From Capacity.", font=LINK)
r = pnl(r, "eff", "Effective tuition per student",
        f"={a('price')}*(1-{a('yr_share')}*{a('yr_disc')})",
        f"=C{r}*{a('months')}", MNT, "After blending the full-year discount across the mix.")
r = pnl(r, "rev", "Tuition revenue", f"=C{P['students']}*C{P['eff']}",
        f"=C{r}*{a('months')}", MNT, "", bold=True)
r = pnl(r, "reg", "Registration / materials fee",
        f"={a('reg_fee')}*C{P['students']}/{a('months')}",
        f"={a('reg_fee')}*C{P['students']}", MNT, "One-off, spread across the year.")
r = pnl(r, "totrev", "TOTAL REVENUE", f"=C{P['rev']}+C{P['reg']}",
        f"=D{P['rev']}+D{P['reg']}", MNT, "", bold=True)
ws.cell(row=P["totrev"], column=3).fill = SEC_FILL
ws.cell(row=P["totrev"], column=4).fill = SEC_FILL

r += 1
r = section(ws, r, "  COSTS")
r = pnl(r, "c_teach", "Hired teachers", f"=-{KO}{K['c_teach']}", f"=C{r}*{a('months')}", MNT, "", font=LINK)
r = pnl(r, "c_own", "Owner teaching time (costed)", f"=-{KO}{K['c_owner']}", f"=C{r}*{a('months')}", MNT, "NOT free — see README.", font=LINK)
r = pnl(r, "c_admin", "Admin", f"=-{KO}{K['c_admin']}", f"=C{r}*{a('months')}", MNT, "", font=LINK)
r = pnl(r, "c_fac", "Rent + utilities", f"=-({KO}{K['c_rent']}+{KO}{K['c_util']})", f"=C{r}*12", MNT, "Rent runs 12 months even if you teach 10.", font=LINK)
r = pnl(r, "c_tech", "Technology", f"=-{KO}{K['c_tech']}", f"=C{r}*12", MNT, "", font=LINK)
r = pnl(r, "c_mkt", "Marketing", f"=-{KO}{K['c_mkt']}", f"=C{r}*12", MNT, "", font=LINK)
r = pnl(r, "c_oth", "Other / contingency", f"=-{KO}{K['c_oth']}", f"=C{r}*12", MNT, "", font=LINK)
r = pnl(r, "c_var", "Materials + food (variable)", f"=-{KO}{K['var']}*C{P['students']}", f"=C{r}*{a('months')}", MNT, "", font=LINK)
r = pnl(r, "totcost", "TOTAL COST", f"=SUM(C{P['c_teach']}:C{P['c_var']})",
        f"=SUM(D{P['c_teach']}:D{P['c_var']})", MNT, "", bold=True)
ws.cell(row=P["totcost"], column=3).fill = SEC_FILL
ws.cell(row=P["totcost"], column=4).fill = SEC_FILL

r += 1
r = section(ws, r, "  RESULT")
r = pnl(r, "profit", "PROFIT", f"=C{P['totrev']}+C{P['totcost']}",
        f"=D{P['totrev']}+D{P['totcost']}", MNT, "", bold=True)
ws.cell(row=P["profit"], column=3).fill = KEY_FILL
ws.cell(row=P["profit"], column=4).fill = KEY_FILL
r = pnl(r, "margin", "Margin", f"=IFERROR(C{P['profit']}/C{P['totrev']},0)",
        f"=IFERROR(D{P['profit']}/D{P['totrev']},0)", PCT, "")
r = pnl(r, "be", "Break-even students",
        f"=IFERROR({KO}{K['fixed']}/(C{P['eff']}-{KO}{K['var']}),\"n/a\")",
        f"=C{r}", NUM0, "At this price.")
r = pnl(r, "cushion", "Seats of cushion above break-even",
        f"=C{P['students']}-C{P['be']}", f"=C{r}", NUM1,
        "How many students you can lose before the centre loses money.")
ws.cell(row=P["cushion"], column=3).fill = KEY_FILL
r += 1
r = note(ws, r, "Profit here is AFTER charging the owner's teaching time at the opportunity-cost rate. A small positive profit therefore still means the owner is being paid for teaching. A negative profit means the centre is not covering the owner's time on top of its bills.", col=2)

PN = "PnL!$C$"

# ============================================================== ENROLMENT
ws = sheet("Enrolment", [3, 44, 15, 15, 15, 15, 52])
r = 1
r = title(ws, r, "Full school year vs monthly subscription")
r = note(ws, r, "Same student, two contracts. The question is what churn costs you, and whether the full-year discount is cheaper than that churn.", col=2)
r += 1

r = section(ws, r, "  ONE STUDENT, ONE SCHOOL YEAR", span=7)
ws.cell(row=r, column=3, value="Full year").font = H3
ws.cell(row=r, column=4, value="Monthly").font = H3
ws.cell(row=r, column=5, value="Difference").font = H3
r += 1
E = {}


def enr(row, key, label, f_year, f_month, src="", fmt=MNT, bold=False):
    ws.cell(row=row, column=2, value=label).font = H3 if bold else BODY
    for col, f in ((3, f_year), (4, f_month)):
        c = ws.cell(row=row, column=col, value=f)
        c.font = Font(name=FONT, size=10, bold=True) if bold else CALC
        c.number_format = fmt
        c.border = BOX
    c = ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
    c.font = Font(name=FONT, size=10, bold=True) if bold else CALC
    c.number_format = fmt
    c.border = BOX
    ws.cell(row=row, column=7, value=src).font = SMALL
    ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="top")
    E[key] = row
    return row + 1


r = enr(r, "headline", "Headline price per month", f"={a('price')}", f"={a('price')}", "Same list price.")
r = enr(r, "disc", "Less full-year discount",
        f"=-{a('price')}*{a('yr_disc')}", "=0", "Discount is the price of commitment.")
r = enr(r, "net", "Net price per month", f"=C{E['headline']}+C{E['disc']}", f"=D{E['headline']}+D{E['disc']}", "")
r = enr(r, "months_paid", "Expected months paid",
        f"={a('months')}",
        f"=IFERROR((1-(1-{a('churn')})^{a('months')})/{a('churn')},{a('months')})", "",
        fmt=NUM1)
ws.cell(row=E["months_paid"], column=7, value="Monthly: expected months before churn, over a 10-month year. Full year: committed.").font = SMALL
r = enr(r, "gross", "Revenue per student per year",
        f"=C{E['net']}*C{E['months_paid']}", f"=D{E['net']}*D{E['months_paid']}", "", bold=True)
r = enr(r, "refill", "Less cost to replace the leaver",
        "=0", f"=-{a('cac')}*(1-(C{E['months_paid']}=D{E['months_paid']}))*{a('churn')}*{a('months')}",
        "Marketing spend to refill an empty seat.")
r = enr(r, "empty", "Less lost margin on the empty seat",
        "=0",
        f"=-(C{E['months_paid']}-D{E['months_paid']})*(D{E['net']}-{KO}{K['var']})*0.5",
        "A seat that empties mid-year is rarely refilled instantly. Assumes half the lost months are recovered.")
r = enr(r, "netrev", "NET CONTRIBUTION PER STUDENT PER YEAR",
        f"=C{E['gross']}+C{E['refill']}+C{E['empty']}",
        f"=D{E['gross']}+D{E['refill']}+D{E['empty']}", "", bold=True)
ws.cell(row=E["netrev"], column=3).fill = KEY_FILL
ws.cell(row=E["netrev"], column=4).fill = KEY_FILL
ws.cell(row=E["netrev"], column=5).fill = KEY_FILL

r += 1
ws.cell(row=r, column=2, value="Is the full-year discount worth it?").font = H3
c = ws.cell(row=r, column=3, value=f'=IF(E{E["netrev"]}>0,"YES — commitment beats churn at these settings","NO — the discount costs more than the churn it prevents")')
c.font = WARN
r += 1
r = note(ws, r, f"The answer flips on two numbers only: the discount (Assumptions) and monthly churn (Assumptions). Find the churn rate at which they break even by changing churn until the difference above reaches zero.", col=2)
r += 1

r = section(ws, r, "  WHOLE CENTRE, AT THE SELECTED MIX", span=7)
r = enr(r, "mix", "Students on each plan",
        f"={CAP}{C['students']}*{a('yr_share')}",
        f"={CAP}{C['students']}*(1-{a('yr_share')})", "", fmt=NUM1)
r = enr(r, "mixrev", "Net contribution per year",
        f"=C{E['mix']}*C{E['netrev']}", f"=D{E['mix']}*D{E['netrev']}", "", bold=True)
r += 1
ws.cell(row=r, column=2, value="Combined net contribution per year").font = H3
c = ws.cell(row=r, column=3, value=f"=C{E['mixrev']}+D{E['mixrev']}")
c.font = Font(name=FONT, size=11, bold=True)
c.number_format = MNT
c.fill = KEY_FILL
r += 2
r = note(ws, r, "Practical note not in the arithmetic: a full-year contract in Mongolia is only as good as the family's ability to pay through February. Tsagaan Sar puts heavy pressure on household cash. A full-year commitment paid in 10 monthly instalments carries most of the retention benefit with far less payment risk than demanding the year upfront — and the CashFlow sheet shows why February is the month to worry about.", col=2)

# ================================================================ WEBSITE
ws = sheet("Website", [3, 48, 17, 17, 15, 56])
W = {}
r = 1
r = title(ws, r, "The website: bundled, add-on, and as its own product")
r += 1

r = section(ws, r, "  QUESTION 1 — BUNDLE IT INTO THE CENTRE FEE, OR SELL IT AS AN ADD-ON?", span=6)
r = note(ws, r, "The strategic argument for bundling is that the site is what makes the high fee legible: it produces the monthly evidence report. Modelled below with numbers, not opinion.", col=2)
r += 1
ws.cell(row=r, column=3, value="Bundled").font = H3
ws.cell(row=r, column=4, value="Add-on").font = H3
ws.cell(row=r, column=5, value="Diff").font = H3
r += 1


def web(row, key, label, f_b, f_a, src="", fmt=MNT, bold=False):
    ws.cell(row=row, column=2, value=label).font = H3 if bold else BODY
    for col, f in ((3, f_b), (4, f_a)):
        c = ws.cell(row=row, column=col, value=f)
        c.font = Font(name=FONT, size=10, bold=True) if bold else CALC
        c.number_format = fmt
        c.border = BOX
    c = ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
    c.font = Font(name=FONT, size=10, bold=True) if bold else CALC
    c.number_format = fmt
    c.border = BOX
    ws.cell(row=row, column=6, value=src).font = SMALL
    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    W[key] = row
    return row + 1


r = web(r, "students", "Centre students", f"={CAP}{C['students']}", f"={CAP}{C['students']}", "Same roll either way.", NUM0)
r = web(r, "tuition", "Tuition per student per month",
        f"={a('price')}*(1+{a('bundle_uplift')})", f"={a('price')}",
        "Bundling supports a higher headline fee because the report justifies it.")
r = web(r, "addon", "Add-on revenue per student per month",
        "=0", f"={a('web_addon')}*{a('web_attach')}",
        "Add-on sold to a fraction of families only.")
r = web(r, "total_pm", "Revenue per student per month",
        f"=C{W['tuition']}+C{W['addon']}", f"=D{W['tuition']}+D{W['addon']}", "", bold=True)
r = web(r, "churn", "Monthly churn",
        f"=MAX(0,{a('churn')}-{a('bundle_churn')})", f"={a('churn')}",
        "Every family sees the report every month, so bundling also retains.", PCT)
r = web(r, "months", "Expected months paid",
        f"=IFERROR((1-(1-C{W['churn']})^{a('months')})/C{W['churn']},{a('months')})",
        f"=IFERROR((1-(1-D{W['churn']})^{a('months')})/D{W['churn']},{a('months')})", "", NUM1)
r = web(r, "annual", "Revenue per student per year",
        f"=C{W['total_pm']}*C{W['months']}", f"=D{W['total_pm']}*D{W['months']}", "", bold=True)
r = web(r, "centre", "CENTRE REVENUE PER YEAR",
        f"=C{W['annual']}*C{W['students']}", f"=D{W['annual']}*D{W['students']}", "", bold=True)
ws.cell(row=W["centre"], column=3).fill = KEY_FILL
ws.cell(row=W["centre"], column=4).fill = KEY_FILL
ws.cell(row=W["centre"], column=5).fill = KEY_FILL
r += 1
ws.cell(row=r, column=2, value="Which wins, on these assumptions?").font = H3
c = ws.cell(row=r, column=3, value=f'=IF(E{W["centre"]}>0,"BUNDLE","ADD-ON")')
c.font = WARN
r += 1
r = note(ws, r, "This result is driven almost entirely by two yellow cells on Assumptions: the tuition uplift the bundle supports, and the churn reduction it delivers. Neither is observed — they are the hypotheses the first year should test. The break-even is: bundling wins whenever (uplift × tuition) exceeds (add-on price × attach rate), before even counting retention.", col=2)
r += 1
r = note(ws, r, "There is also a non-financial argument the spreadsheet cannot hold: an add-on splits the parent body into those who can see the evidence and those who cannot. If the report is the proof of the premium, then charging separately for the proof undermines the premium.", col=2)
r += 1

r = section(ws, r, "  QUESTION 2 — WHAT IS THE VERIFIED-ATTEMPT ADVANTAGE WORTH?", span=6)
for t in [
    "Mechanism (as described in the brief — doc 05 could not be read, see README): attempts made in the room are supervised. A human",
    "watched them happen. That makes the predicted ЭЕШ score trustworthy in a way no online-only competitor in Mongolia can match,",
    "because their data cannot distinguish a student's own work from a phone under the desk.",
    "",
    "Expressed to a parent, the product is not 'lessons'. It is: a defensible monthly answer to 'how is my child actually doing, and what",
    "happens if nothing changes?' — with the score projection, the topic-level gaps, and the evidence it rests on.",
]:
    r = note(ws, r, t, col=2)
r += 1
ws.cell(row=r, column=2, value="Converting that into MNT — what a price premium is worth per year").font = H3
r += 1
for i, h in enumerate(["", "Premium over the visible competitor", "Annual fee/student", "Extra revenue/student/yr", "Centre extra revenue/yr", ""], start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.fill = HDR_FILL
    c.font = H2
    c.alignment = Alignment(wrap_text=True, vertical="center")
r += 1
base_row = r
for label, mult in [("At parity with Eduzone (1.0×)", 1.0), ("1.5× the visible competitor", 1.5),
                    ("2.0×", 2.0), ("2.5×", 2.5), ("3.0×", 3.0), ("4.0×", 4.0)]:
    ws.cell(row=r, column=2, value=label).font = BODY
    ws.cell(row=r, column=3, value=mult).font = INPUT
    ws.cell(row=r, column=3).number_format = '0.0"×"'
    ws.cell(row=r, column=4, value=f"=C{r}*1800000").number_format = MNT
    ws.cell(row=r, column=4).font = CALC
    ws.cell(row=r, column=5, value=f"=D{r}-1800000").number_format = MNT
    ws.cell(row=r, column=5).font = CALC
    ws.cell(row=r, column=6, value=f"=E{r}*{CAP}{C['students']}").number_format = MNT
    ws.cell(row=r, column=6).font = CALC
    for col in range(2, 7):
        ws.cell(row=r, column=col).border = BOX
    r += 1
r += 1
r = note(ws, r, "Read this as: 'to justify charging 2.5× what Eduzone charges, the monthly report has to be worth about 2.7m₮ a year to a parent.' That is the sentence to test on ten real parents before launch. Note the anchor: UB families already pay 8–10m₮/year for private school, and ISU families pay $24–44k — so the ceiling on willingness to pay is set by perceived quality, not by the tutoring market.", col=2)
r += 1

r = section(ws, r, "  QUESTION 3 — THE PUBLIC WEBSITE AS ITS OWN PRODUCT", span=6)
r = note(ws, r, "Non-centre users. Twelve registered today, no monetisation. Modelled as a funnel.", col=2)
r += 1
PUB = {}
r = calc(r, "u12", "Registered users after 12 months",
         f"={a('pub_users')}+{a('pub_growth')}*12", NUM0, "Growth is an assumption — no traffic data was supplied.", PUB)
r = calc(r, "paid", "Paying users at that point",
         f"=C{PUB['u12']}*{a('pub_conv')}", NUM1, "", PUB)
r = calc(r, "mrr", "Monthly revenue",
         f"=C{PUB['paid']}*{a('pub_price')}", MNT, "", PUB)
r = calc(r, "arr", "Annual revenue (run rate)",
         f"=C{PUB['mrr']}*12", MNT, "", PUB)
r = calc(r, "vs", "…as a share of centre revenue",
         f"=IFERROR(C{PUB['arr']}/{PN}{P['totrev']}/{a('months')},0)", PCT,
         "Sense-check on where the effort should go.", PUB)
ws.cell(row=PUB["vs"], column=3).fill = KEY_FILL
r += 1
r = note(ws, r, "On any plausible assumption the public site is a rounding error next to the centre in year one. That is not an argument against building it — it is an argument for treating it as (a) the marketing funnel that fills the centre and (b) the delivery mechanism for the parent report, and NOT as a revenue line that deserves to compete for attention with filling classrooms.", col=2)
r += 1
ws.cell(row=r, column=2, value="Suggested tier split — to be decided, not modelled").font = H3
r += 1
for t in [
    "FREE  — practice questions, worked solutions, unverified self-scored progress. Enough to be genuinely useful and to rank in search.",
    "PAID  — full-length timed papers, the score projection, topic-level gap analysis, and history over time.",
    "CENTRE ONLY — the VERIFIED score projection and the monthly parent report. This is the thing that cannot be bought online at any price,",
    "        because it requires a supervised room. Keeping it centre-only is what stops the website cannibalising the centre.",
]:
    r = note(ws, r, t, col=2)

# =============================================================== CASHFLOW
ws = sheet("CashFlow", [3, 30, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11])
r = 1
r = title(ws, r, "Cash-flow calendar — Aug 2026 to Aug 2027")
r = note(ws, r, "Against the Mongolian school year. Student numbers in blue are editable — they are a judgement, not an output.", col=2)
r += 1

months = ["Aug-26", "Sep-26", "Oct-26", "Nov-26", "Dec-26", "Jan-27", "Feb-27",
          "Mar-27", "Apr-27", "May-27", "Jun-27", "Jul-27", "Aug-27"]
ws.cell(row=r, column=2, value="").font = BODY
for i, m in enumerate(months):
    c = ws.cell(row=r, column=3 + i, value=m)
    c.fill = HDR_FILL
    c.font = H2
    c.alignment = Alignment(horizontal="center")
hdr_row = r
r += 1

# phase labels
phases = ["Setup", "Start", "Report", "Peak", "Peak", "Intake", "Tsagaan Sar",
          "Peak", "Peak", "Exam run", "ЭЕШ", "Summer", "Setup"]
ws.cell(row=r, column=2, value="Phase").font = SMALL
for i, p in enumerate(phases):
    c = ws.cell(row=r, column=3 + i, value=p)
    c.font = SMALL
    c.alignment = Alignment(horizontal="center")
r += 1

# students by month — editable
stu_row = r
ws.cell(row=r, column=2, value="Students enrolled").font = H3
default_students = [0, 0.55, 0.75, 0.90, 0.90, 1.00, 0.95, 0.95, 0.95, 0.95, 0.90, 0.25, 0.10]
for i, frac in enumerate(default_students):
    c = ws.cell(row=r, column=3 + i, value=f"=ROUND({CAP}{C['students']}*{frac},0)")
    c.font = INPUT
    c.number_format = NUM0
    c.border = BOX
    c.fill = KEY_FILL
r += 1
ws.cell(row=r, column=2, value="  (as % of steady-state roll)").font = SMALL
for i, frac in enumerate(default_students):
    c = ws.cell(row=r, column=3 + i, value=frac)
    c.font = SMALL
    c.number_format = PCT
    c.alignment = Alignment(horizontal="center")
r += 2

cf_first = r
ws.cell(row=r, column=2, value="Tuition revenue").font = BODY
for i in range(13):
    col = get_column_letter(3 + i)
    teaching = 0 if i in (0, 11, 12) else 1
    c = ws.cell(row=r, column=3 + i,
                value=f"={col}{stu_row}*{a('price')}*(1-{a('yr_share')}*{a('yr_disc')})*{teaching}")
    c.font = CALC
    c.number_format = MNT
rev_row = r
r += 1

ws.cell(row=r, column=2, value="Fixed costs").font = BODY
always = (f"{KO}{K['c_admin']}+{KO}{K['c_rent']}+{KO}{K['c_util']}+{KO}{K['c_tech']}"
          f"+{KO}{K['c_mkt']}+{KO}{K['c_oth']}")
for i in range(13):
    teaching = 0 if i in (0, 11, 12) else 1
    c = ws.cell(row=r, column=3 + i,
                value=f"=-({always})-{KO}{K['c_teach']}*MAX({teaching},{a('pay_summer')})"
                      f"-{KO}{K['c_owner']}*{teaching}")
    c.font = CALC
    c.number_format = MNT
fix_row = r
r += 1
ws.cell(row=r, column=2, value="  (owner time not charged in non-teaching months)").font = SMALL
r += 1

ws.cell(row=r, column=2, value="Variable costs").font = BODY
for i in range(13):
    col = get_column_letter(3 + i)
    teaching = 0 if i in (0, 11, 12) else 1
    c = ws.cell(row=r, column=3 + i, value=f"=-{col}{stu_row}*{KO}{K['var']}*{teaching}")
    c.font = CALC
    c.number_format = MNT
var_row = r
r += 1

ws.cell(row=r, column=2, value="One-off setup / launch marketing").font = BODY
setup = [-4000000, -1000000, 0, 0, 0, -1000000, 0, 0, 0, 0, 0, 0, -3000000]
for i, v in enumerate(setup):
    c = ws.cell(row=r, column=3 + i, value=v)
    c.font = INPUT
    c.number_format = MNT
    c.border = BOX
setup_row = r
r += 1
ws.cell(row=setup_row, column=2).comment = None

ws.cell(row=r, column=2, value="NET CASH").font = H3
for i in range(13):
    col = get_column_letter(3 + i)
    c = ws.cell(row=r, column=3 + i,
                value=f"={col}{rev_row}+{col}{fix_row}+{col}{var_row}+{col}{setup_row}")
    c.font = Font(name=FONT, size=10, bold=True)
    c.number_format = MNT
net_row = r
r += 1

ws.cell(row=r, column=2, value="CUMULATIVE CASH").font = H3
for i in range(13):
    col = get_column_letter(3 + i)
    prev = get_column_letter(2 + i)
    f = f"={col}{net_row}" if i == 0 else f"={prev}{r}+{col}{net_row}"
    c = ws.cell(row=r, column=3 + i, value=f)
    c.font = Font(name=FONT, size=10, bold=True)
    c.number_format = MNT
cum_row = r
r += 2

for row_ in (net_row, cum_row):
    ws.conditional_formatting.add(
        f"C{row_}:O{row_}", CellIsRule(operator="lessThan", formula=["0"], fill=red))
    ws.conditional_formatting.add(
        f"C{row_}:O{row_}", CellIsRule(operator="greaterThan", formula=["0"], fill=green))

r = section(ws, r, "  THE TIGHT MONTHS", span=15)
for t in [
    "AUG 2026 — the worst month, and it is now. Fit-out, furniture, deposits and launch marketing all land before a single tugrik of tuition.",
    "   Rent, utilities and any hired teacher's salary start before the first lesson. This is the month that needs cash in the bank, not a forecast.",
    "",
    "FEB 2027 — Tsagaan Sar. Household cash goes to family obligations; tuition is exactly the kind of payment that slips. Expect late payment",
    "   rather than cancellation, but model it: this is the month a monthly-subscription roll quietly shrinks. (Confirm the 2027 dates.)",
    "",
    "JUL–AUG 2027 — the post-exam cliff. ЭЕШ is sat in late June; the graduating cohort leaves at once and does not come back. Fixed costs —",
    "   rent above all — continue through a near-zero revenue summer, straight into next year's setup spend. Two consecutive drains.",
    "",
    "The structural point: this business collects revenue over ten months and pays rent over twelve. A summer intensive, or a deliberate",
    "   cash reserve carried out of the spring, is not optional — it is what makes September possible.",
]:
    r = note(ws, r, t, col=2)

# =============================================================== SCENARIOS
ws = sheet("Scenarios", [3, 44, 18, 18, 18, 50])
r = 1
r = title(ws, r, "Scenarios")
r = note(ws, r, "Three coherent pictures. Edit the blue inputs in each column. Costs come from the Costs sheet in every case.", col=2)
r += 1

for i, h in enumerate(["", "", "Conservative", "Base", "Stretch", ""], start=1):
    c = ws.cell(row=r, column=i, value=h)
    c.fill = HDR_FILL
    c.font = H2
    c.alignment = Alignment(horizontal="center")
r += 1
S = {}


def scen(row, key, label, vals, fmt, src="", is_input=True):
    ws.cell(row=row, column=2, value=label).font = BODY
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=3 + j, value=v)
        c.font = INPUT if is_input else CALC
        c.number_format = fmt
        c.border = BOX
        if is_input:
            c.fill = KEY_FILL
    ws.cell(row=row, column=6, value=src).font = SMALL
    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    S[key] = row
    return row + 1


r = scen(r, "price", "Tuition ₮/student/month", [350000, 450000, 650000], MNT, "The lever you are choosing.")
r = scen(r, "cohort", "Cohort size", [12, 16, 20], NUM0, "Bigger cohorts earn more but dilute the premium.")
r = scen(r, "fill", "Seat fill rate", [0.65, 0.85, 0.95], PCT, "")
r = scen(r, "util", "Slot utilisation", [0.45, 0.55, 0.70], PCT, "")
r += 1

for key, label, formula in [
    ("slots", "Realistic slots/week/room",
     "=(({wd}*{wdd})+({we}*{wed}))*{{u}}"),
]:
    pass

# derived rows
def sderive(row, key, label, f_tpl, fmt, src="", bold=False):
    ws.cell(row=row, column=2, value=label).font = H3 if bold else BODY
    for j, col in enumerate("CDE"):
        c = ws.cell(row=row, column=3 + j, value=f_tpl.format(c=col))
        c.font = Font(name=FONT, size=10, bold=True) if bold else CALC
        c.number_format = fmt
        c.border = BOX
    ws.cell(row=row, column=6, value=src).font = SMALL
    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    S[key] = row
    return row + 1


slots_expr = f"((({a('wd_slots')}*{a('wd_days')})+({a('we_slots')}*{a('we_days')}))*{{c}}{S['util']})"
cohorts_expr = f"({slots_expr}/{a('sess_wk')})"
students_expr_s = (f"ROUND({cohorts_expr}*MIN({{c}}{S['cohort']},{a('capA')})*{{c}}{S['fill']}"
                   f"+{cohorts_expr}*MIN({a('cohortB')},{a('capB')})*{{c}}{S['fill']}*{a('useB')},0)")

r = sderive(r, "students", "Students", f"={students_expr_s}", NUM0, "")
r = sderive(r, "hours", "Contact hours/week", f"={slots_expr}*(1+{a('useB')})*{a('sess_hours')}", NUM1, "")
r = sderive(r, "teachers", "Hired teachers required",
            f"=ROUNDUP(MAX(0,{slots_expr}*(1+{a('useB')})*{a('sess_hours')}-{a('own_hrs')})/{a('teach_hrs')},0)", NUM0, "")
r = sderive(r, "fixed", "Fixed cost / month",
            f"={{c}}{S['teachers']}*{a('teach_sal')}*(1+{a('si')})"
            f"+{a('own_hrs')}*{a('weeks')}*{a('own_rate')}"
            f"+{a('admin_n')}*{a('admin_sal')}*(1+{a('si')})"
            f"+{a('rent')}+{a('util_cost')}+{a('internet')}"
            f"+({a('supabase')}+{a('hosting')})*{a('fx')}+{a('marketing')}+{a('other')}", MNT,
            "Same cost logic as the Costs sheet, recomputed per scenario.")
r = sderive(r, "rev", "Revenue / month",
            f"={{c}}{S['students']}*{{c}}{S['price']}*(1-{a('yr_share')}*{a('yr_disc')})", MNT, "")
r = sderive(r, "var", "Variable cost / month",
            f"={{c}}{S['students']}*{KO}{K['var']}", MNT, "")
r = sderive(r, "profit", "PROFIT / MONTH",
            f"={{c}}{S['rev']}-{{c}}{S['var']}-{{c}}{S['fixed']}", MNT, "", bold=True)
r = sderive(r, "profity", "PROFIT / YEAR",
            f"={{c}}{S['profit']}*{a('months')}", MNT,
            "After paying the owner for teaching. Rent in the two non-teaching months is not deducted here — see CashFlow.", bold=True)
r = sderive(r, "be", "Break-even students",
            f"=IFERROR({{c}}{S['fixed']}/({{c}}{S['price']}*(1-{a('yr_share')}*{a('yr_disc')})-{KO}{K['var']}),\"n/a\")", NUM0, "")
r = sderive(r, "cush", "Cushion (students above break-even)",
            f"={{c}}{S['students']}-{{c}}{S['be']}", NUM1,
            "Negative means this scenario loses money.", bold=True)
r = sderive(r, "vsedu", "Annual fee vs Eduzone (1.8m₮)",
            f"={{c}}{S['price']}*{a('months')}/1800000", '0.0"×"',
            "How big an ask this is, against the only clearly published competitor price.")
r = sderive(r, "concur", "Students in building, parallel session",
            f"=MIN({{c}}{S['cohort']},{a('capA')})+MIN({a('cohortB')},{a('capB')})*{a('useB')}", NUM0, "")
r = sderive(r, "capchk", "Within the stated ~20 capacity?",
            f'=IF({{c}}{S["concur"]}<={a("concur_cap")},"OK","OVER")', NUM0,
            "Physical check on the flat, scenario by scenario.")
r = sderive(r, "phys", "Physical ceiling on the roll (100% fill)",
            f"=((({a('wd_slots')}*{a('wd_days')})+({a('we_slots')}*{a('we_days')}))*{{c}}{S['util']}"
            f"*({a('capA')}+{a('capB')}*{a('useB')}))/{a('sess_wk')}", NUM0,
            "The roll above must sit below this or the timetable does not exist.")

for key in ("profit", "profity", "cush"):
    for col in "CDE":
        ws.conditional_formatting.add(f"{col}{S[key]}:{col}{S[key]}",
                                      CellIsRule(operator="lessThan", formula=["0"], fill=red))
        ws.conditional_formatting.add(f"{col}{S[key]}:{col}{S[key]}",
                                      CellIsRule(operator="greaterThan", formula=["0"], fill=green))
r += 1
r = note(ws, r, "Note what the Conservative column is really testing: it is not 'a bad year', it is 'the premium did not land'. Lower price, smaller cohorts, emptier seats. If that column survives, the downside is bounded. If it does not, the plan depends on the premium working first time — which is a different kind of risk and worth knowing before September.", col=2)
r += 1
r = note(ws, r, "TREAT STRETCH WITH SUSPICION. It compounds a high price, the largest cohort, near-full seats AND high utilisation at the same time. Each is individually possible; all four together, in year one, in a converted flat, is not a plan — it is a coincidence. The Base column is the one to argue about.", col=2)

del wb["Sheet"]
wb.save("/home/user/mathematica/outputs/pricing/MongolPotential-Pricing-Model.xlsx")
print("saved")
